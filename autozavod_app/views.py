from django.shortcuts import render, get_object_or_404, redirect
from .models import Stuff, EVP_CLASS, Doc, Process, Action, TypeDoc, Workshop, Orders
from django.http import HttpResponse
from django.urls import reverse
from .forms import WorkshopForm, DocForm, OrdersForm, ProcessForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.db.models import Q
import openpyxl
import pandas as pd
import json


def show_menu(request):
    if request.method == "POST":
        form = WorkshopForm(request.POST)
        if form.is_valid():

            request.session["workshop"] = form.cleaned_data["workshop"].id
            selected = form.cleaned_data["workshop"]

            return redirect(f"/docs/{selected.workshop_name}")
    else:
        form = WorkshopForm()
        return render(request, "autozavod_app/show_menu.html", {"form": form})


def show_docs(request, workshop_name):
    workshop_id = request.session.get("workshop")

    if not workshop_id:
        return HttpResponse("Цех не выбран или ошибка")

    workshop = get_object_or_404(Workshop, id=workshop_id)
    docs = Doc.objects.filter(workshop=workshop)

    # Получаем параметр сортировки из запроса
    sort_by = request.GET.get(
        "sort_by", "date"
    )  # По умолчанию сортируем по дате выпуска

    # Сортируем документы в зависимости от параметра
    if sort_by == "name":
        docs = docs.order_by("name")  # Сортировка по алфавиту
    elif sort_by == "release_date":
        docs = docs.order_by("-date")  # Сортировка по новизне (сначала новые)
    elif sort_by == "date":
        docs = docs.order_by("date")  # Сортировка по дате (сначала старые)

    # Возвращаем шаблон с данными
    return render(
        request,
        "autozavod_app/show_docs.html",
        {"docs": docs, "workshop": workshop, "sort_by": sort_by},
    )


def show_orders(request, workshop_name):

    workshop_id = request.session.get("workshop")

    sort_by = request.GET.get(
        "sort_by", "release_date"
    )  # По умолчанию сортируем по дате выпуска

    # Сортируем распоряжения в зависимости от параметра
    if sort_by == "name":
        orders = Orders.objects.all().order_by("name")  # Сортировка по алфавиту
    elif sort_by == "release_date":
        orders = Orders.objects.all().order_by(
            "-release_date"
        )  # Сортировка по новизне (сначала новые)
    elif sort_by == "date":
        orders = Orders.objects.all().order_by(
            "release_date"
        )  # Сортировка по дате (сначала старые)
    else:
        orders = (
            Orders.objects.all()
        )  # Если параметр не указан, возвращаем без сортировки

    if not workshop_id:
        return HttpResponse("Цех не выбран или ошибка")
    workshop = get_object_or_404(Workshop, id=workshop_id)
    orders = Orders.objects.filter(workshop=workshop)

    return render(
        request,
        "autozavod_app/show_all_orders.html",
        {"order": orders, "workshop": workshop, "sort_by": sort_by},
    )


def show_all_orders(request):
    # Получаем параметр сортировки из запроса
    sort_by = request.GET.get(
        "sort_by", "release_date"
    )  # По умолчанию сортируем по дате выпуска

    # Сортируем распоряжения в зависимости от параметра
    if sort_by == "name":
        orders = Orders.objects.all().order_by("name")  # Сортировка по алфавиту
    elif sort_by == "release_date":
        orders = Orders.objects.all().order_by(
            "-release_date"
        )  # Сортировка по новизне (сначала новые)
    elif sort_by == "date":
        orders = Orders.objects.all().order_by(
            "release_date"
        )  # Сортировка по дате (сначала старые)
    else:
        orders = (
            Orders.objects.all()
        )  # Если параметр не указан, возвращаем без сортировки

    # Возвращаем шаблон с данными
    return render(
        request,
        "autozavod_app/show_all_orders.html",
        {
            "order": orders,
            "sort_by": sort_by,  # Передаём параметр сортировки в шаблон
        },
    )


def show_all_docs(request):

    docs = Doc.objects.prefetch_related("workshop").all()
    # Получаем параметр сортировки из запроса
    sort_by = request.GET.get(
        "sort_by", "date"
    )  # По умолчанию сортируем по дате выпуска

    # Сортируем документы в зависимости от параметра
    if sort_by == "name":
        docs = docs.order_by("name")  # Сортировка по алфавиту
    elif sort_by == "release_date":
        docs = docs.order_by("date")  # Сортировка по новизне (сначала новые)
    elif sort_by == "date":
        docs = docs.order_by("-date")  # Сортировка по дате (сначала старые)

    return render(
        request, "autozavod_app/show_all_docs.html", {"docs": docs, "sort_by": sort_by}
    )


def doc_show_info(request, id):
    doc = get_object_or_404(Doc, id=id)
    doc.formatted_date = doc.date.strftime("%d.%m.%Y")

    # Получаем все процессы, связанные с документом
    processes = doc.process.all()

    # Проверяем, все ли процессы завершены (status_procces == True)
    all_completed_process = all(process.status_procces is True for process in processes)

    # Получаем все связанные распоряжения (Orders), в которых есть эти процессы
    orders = doc.orders.all()

    # Проверяем, выполнены ли все процессы во всех Orders
    all_completed_process_orders = True
    for order in orders:
        for process in order.process.all():
            if not process.status_procces:
                all_completed_process_orders = False
                break
        if not all_completed_process_orders:
            break

    # Проверяем, есть ли хотя бы один Order, где есть не ознакомленные сотрудники
    any_not_acquainted = False
    for order in orders:
        if order.responsible_not.exists():
            any_not_acquainted = True
            break

    # Все ознакомились, если в ни одном распоряжении нет неознакомленных сотрудников
    all_responsible_acquainted_orders = not any_not_acquainted

    return render(
        request,
        "autozavod_app/doc_show_info.html",
        {
            "doc": doc,
            "all_completed_process": all_completed_process,
            "all_completed_process_orders": all_completed_process_orders,
            "all_responsible_acquainted_orders": all_responsible_acquainted_orders,
            "orders": orders,  # Передаём все распоряжения в шаблон
        },
    )


def create_doc(request):
    if request.method == "POST":
        form = DocForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("success")
    else:
        form = DocForm()
        employees = Stuff.objects.all()  # Получаем всех сотрудников
    return render(
        request, "autozavod_app/create_doc.html", {"employees": employees, "form": form}
    )


def create_order(request):
    if request.method == "POST":
        form = OrdersForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("success")
    else:
        form = OrdersForm()

    return render(request, "autozavod_app/create_order.html", {"form": form})


def success(request):
    return render(
        request,
        "autozavod_app/success.html",
    )


def order_detail(request, id):
    order = get_object_or_404(Orders, id=id)

    processes = order.process.all()

    all_completed_process = all(process.status_procces is True for process in processes)

    return render(
        request,
        "autozavod_app/order_detail.html",
        {"order": order, "all_completed_process": all_completed_process},
    )


@csrf_exempt
def acquaint_employee(request, employee_id, doc_id):
    if request.method == "POST":
        try:
            doc = Doc.objects.get(id=doc_id)
            employee = Stuff.objects.get(id=employee_id)

            doc.responsible_not.remove(employee)
            doc.responsible.add(employee)
            return JsonResponse({"success": True})
        except (Doc.DoesNotExist, Stuff.DoesNotExist):
            return JsonResponse(
                {"success": False, "error": "Document or employee not found"},
                status=404,
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def unacquaint_employee(request, employee_id, doc_id):
    if request.method == "POST":
        try:
            doc = Doc.objects.get(id=doc_id)
            employee = Stuff.objects.get(id=employee_id)

            doc.responsible.remove(employee)
            doc.responsible_not.add(employee)
            return JsonResponse({"success": True})
        except (Doc.DoesNotExist, Stuff.DoesNotExist):
            return JsonResponse(
                {"success": False, "error": "Document or employee not found"},
                status=404,
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def acquaint_all(request, doc_id):
    if request.method == "POST":
        try:
            doc = Doc.objects.get(id=doc_id)
            employees = doc.responsible_not.all()

            for employee in employees:
                doc.responsible.add(employee)
            doc.responsible_not.clear()
            return JsonResponse({"success": True})
        except Doc.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Document not found"}, status=404
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def unacquaint_all(request, doc_id):
    if request.method == "POST":
        try:
            doc = Doc.objects.get(id=doc_id)
            employees = doc.responsible.all()

            for employee in employees:
                doc.responsible_not.add(employee)
            doc.responsible.clear()
            return JsonResponse({"success": True})
        except Doc.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Document not found"}, status=404
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


def add_process(request, doc_id):
    doc = get_object_or_404(Doc, id=doc_id)

    if request.method == "POST":
        form = ProcessForm(request.POST, request.FILES)
        if form.is_valid():
            process = form.save(commit=False)
            process.save()
            form.save_m2m()
            doc.process.add(process)
            return redirect("doc_show_info", id=doc.id)
    else:
        form = ProcessForm()

    return render(
        request, "autozavod_app/create_process.html", {"form": form, "doc": doc}
    )


def add_order(request, doc_id):
    doc = get_object_or_404(Doc, id=doc_id)

    if request.method == "POST":
        form = OrdersForm(request.POST, request.FILES)
        if form.is_valid():
            order = form.save(commit=False)
            order.save()
            form.save_m2m()
            doc.orders.add(order)
            return redirect("doc_show_info", id=doc.id)
    else:
        form = OrdersForm()

    return render(request, "autozavod_app/add_order.html", {"form": form, "doc": doc})


def add_process_order(request, doc_id):
    doc = get_object_or_404(Orders, id=doc_id)

    if request.method == "POST":
        form = ProcessForm(request.POST, request.FILES)
        if form.is_valid():
            process = form.save(commit=False)
            process.save()
            form.save_m2m()
            doc.process.add(process)
            return redirect("order_detail", id=doc.id)
    else:
        form = ProcessForm()

    return render(
        request, "autozavod_app/create_process.html", {"form": form, "doc": doc}
    )


@require_POST
@csrf_exempt
def delete_process(request, process_id):
    try:
        process = Process.objects.get(id=process_id)
        process.delete()
        print(f"Процесс {process_id} удален")
        return JsonResponse({"status": "success"})
    except Process.DoesNotExist:
        print(f"Процесс {process_id} не найден")
        return JsonResponse(
            {"status": "error", "message": "Процесс не найден"}, status=404
        )


@csrf_exempt
def update_process_status(request, process_id):
    if request.method == "POST":
        try:
            process = Process.objects.get(id=process_id)
            data = json.loads(request.body)
            process.status_procces = data.get("status_procces")
            process.save()
            return JsonResponse({"status": "success"})
        except Process.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Process not found"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})


@csrf_exempt
def acquaint_employee_order(request, employee_id, order_id):
    if request.method == "POST":
        try:
            order = Orders.objects.get(id=order_id)
            employee = Stuff.objects.get(id=employee_id)

            order.responsible_not.remove(employee)
            order.responsible.add(employee)
            return JsonResponse({"success": True})
        except (Orders.DoesNotExist, Stuff.DoesNotExist):
            return JsonResponse(
                {"success": False, "error": "Order or employee not found"},
                status=404,
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def unacquaint_employee_order(request, employee_id, order_id):
    if request.method == "POST":
        try:
            order = Orders.objects.get(id=order_id)
            employee = Stuff.objects.get(id=employee_id)

            order.responsible.remove(employee)
            order.responsible_not.add(employee)
            return JsonResponse({"success": True})
        except (Orders.DoesNotExist, Stuff.DoesNotExist):
            return JsonResponse(
                {"success": False, "error": "Order or employee not found"},
                status=404,
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def acquaint_all_order(request, order_id):
    if request.method == "POST":
        try:
            order = Orders.objects.get(id=order_id)
            employees = order.responsible_not.all()

            for employee in employees:
                order.responsible.add(employee)
            order.responsible_not.clear()
            return JsonResponse({"success": True})
        except Orders.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Order not found"}, status=404
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


@csrf_exempt
def unacquaint_all_order(request, order_id):
    if request.method == "POST":
        try:
            order = Orders.objects.get(id=order_id)
            employees = order.responsible.all()

            for employee in employees:
                order.responsible_not.add(employee)
            order.responsible.clear()
            return JsonResponse({"success": True})
        except Orders.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Order not found"}, status=404
            )
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=400
    )


def report(request):
    if request.method == "POST":
        # Для отчета по сотруднику не нужна форма WorkshopForm
        report_type = request.POST.get("report_type")

        if report_type == "employee":
            employee_id = request.POST.get("employee")
            if employee_id:
                return generate_employee_docs_report(employee_id)
            else:
                # Возвращаем форму с ошибкой
                form = WorkshopForm()
                employees = Stuff.objects.all()
                return render(
                    request,
                    "autozavod_app/report.html",
                    {
                        "form": form,
                        "employees": employees,
                        "error": "Необходимо выбрать сотрудника",
                    },
                )
        else:
            # Для других отчетов используем WorkshopForm
            form = WorkshopForm(request.POST)
            if form.is_valid():
                workshop = form.cleaned_data["workshop"]

                if report_type == "not_responsible":
                    return generate_not_responsible_report(workshop)
                elif report_type == "doc":
                    return generate_doc_report(workshop)
                elif report_type == "order":
                    return generate_order_report(workshop)
                elif report_type == "all":
                    return generate_full_report(workshop)
                elif report_type == "non_actual":
                    return generate_non_actual_docs_report(workshop)

    # GET запрос или невалидная форма
    form = WorkshopForm()
    employees = Stuff.objects.all()
    return render(
        request, "autozavod_app/report.html", {"form": form, "employees": employees}
    )


def generate_not_responsible_report(workshop):
    """Генерация отчёта по сотрудникам, не ознакомившимся с документами и распоряжениями."""
    wb = Workbook()

    # Первый лист: Не ознакомившиеся сотрудники по приказам (Doc)
    ws1 = wb.active
    ws1.title = "Не ознакомившиеся сотрудники по приказам"
    ws1.append(
        ["ФИО сотрудника", "Документ", "ID документа", "Цех", "Участок", "В отпуске"]
    )

    docs = Doc.objects.filter(workshop=workshop)
    for doc in docs:
        if doc.responsible_not.exists():
            for employee in doc.responsible_not.all():
                ws1.append(
                    [
                        employee.name,
                        doc.name,
                        doc.id_doc,
                        employee.workshop.workshop_name if employee.workshop else "",
                        employee.section.section if employee.section else "",
                        employee.on_leave if employee.on_leave else "Нет",
                    ]
                )

    # Второй лист: Не ознакомившиеся сотрудники по распоряжениям (Orders)
    ws2 = wb.create_sheet("Не ознакомившиеся сотрудники по распоряжениям")
    ws2.append(
        [
            "ФИО сотрудника",
            "Распоряжение",
            "ID распоряжения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    orders = Orders.objects.filter(workshop=workshop)
    for order in orders:
        if order.responsible_not.exists():
            for employee in order.responsible_not.all():
                ws2.append(
                    [
                        employee.name,
                        order.name,
                        order.id_doc,
                        employee.workshop.workshop_name if employee.workshop else "",
                        employee.section.section if employee.section else "",
                        employee.on_leave if employee.on_leave else "Нет",
                    ]
                )

    return create_excel_response(wb, "not_responsible_report.xlsx")


def generate_doc_report(workshop):
    """Генерация отчёта по невыполненным действиям в приказах."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Невыполненные приказы"
    ws.append(
        [
            "ФИО сотрудника",
            "Действие",
            "Документ",
            "ID документа",
            "Срок выполнения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    docs = Doc.objects.filter(workshop=workshop)
    for doc in docs:
        processes = doc.process.filter(status_procces=False)
        for process in processes:
            for proc in process.action.all():
                for action in process.responsible_process.all():
                    ws.append(
                        [
                            action.name,
                            proc.action_name,
                            doc.name,
                            doc.id_doc,
                            (
                                process.date_deadline.strftime("%d.%m.%Y")
                                if process.date_deadline
                                else ""
                            ),
                            action.workshop.workshop_name if action.workshop else "",
                            action.section.section if action.section else "",
                            action.on_leave if action.on_leave else "Нет",
                        ]
                    )

    return create_excel_response(wb, "doc_report.xlsx")


def generate_order_report(workshop):
    """Генерация отчёта по невыполненным действиям в распоряжениях."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Невыполненные распоряжения"
    ws.append(
        [
            "Распоряжение",
            "ID распоряжения",
            "Действие",
            "ФИО сотрудника",
            "Срок выполнения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    orders = Orders.objects.filter(workshop=workshop)
    for order in orders:
        processes = order.process.filter(status_procces=False)
        for process in processes:
            for action in process.action.all():
                for responsible in process.responsible_process.all():
                    ws.append(
                        [
                            order.name,
                            order.id_doc,
                            action.action_name,
                            responsible.name,
                            (
                                process.date_deadline.strftime("%d.%m.%Y")
                                if process.date_deadline
                                else ""
                            ),
                            (
                                responsible.workshop.workshop_name
                                if responsible.workshop
                                else ""
                            ),
                            responsible.section.section if responsible.section else "",
                            responsible.on_leave if responsible.on_leave else "Нет",
                        ]
                    )

    return create_excel_response(wb, "order_report.xlsx")


def generate_non_actual_docs_report(workshop):
    """Генерация отчёта по неактуальным документам"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Неактуальные документы"
    ws.append(
        [
            "Название документа",
            "ID документа",
            "Тип документа",
            "Ответственные",
            "Причина неактуальности",
            "Дата выпуска",
        ]
    )

    non_actual_docs = Doc.objects.filter(workshop=workshop, actual="N")
    for doc in non_actual_docs:
        # Определяем причину неактуальности
        reason = "Не указана"
        if doc.mayor.filter(on_leave__isnull=False).exists():
            reason = "Ответственный в отпуске"

        ws.append(
            [
                doc.name,
                doc.id_doc,
                ", ".join([t.type_name for t in doc.type_doc.all()]),
                ", ".join([m.name for m in doc.mayor.all()]),
                reason,
                doc.date.strftime("%d.%m.%Y") if doc.date else "",
            ]
        )

    return create_excel_response(wb, "non_actual_docs.xlsx")


def generate_employee_docs_report(employee_id):
    """Генерация отчёта по всем документам и распоряжениям сотрудника"""
    try:
        employee = Stuff.objects.get(pk=employee_id)
    except Stuff.DoesNotExist:
        return HttpResponse("Сотрудник не найден", status=404)

    wb = Workbook()

    # Лист 1: Ответственный за документы и распоряжения
    ws1 = wb.active
    ws1.title = "Ответственный"
    headers = ["Тип", "Название", "ID", "Тип/Направление", "Статус", "Актуальность", "Дата выпуска"]
    ws1.append(headers)

    # Документы (Doc) где сотрудник ответственный
    for doc in Doc.objects.filter(mayor=employee):
        ws1.append([
            "Документ",
            doc.name,
            doc.id_doc,
            ", ".join([t.type_name for t in doc.type_doc.all()]),
            doc.get_status_display(),
            doc.get_actual_display(),
            doc.date.strftime('%d.%m.%Y') if doc.date else ""
        ])

    # Распоряжения (Orders) где сотрудник ответственный
    for order in Orders.objects.filter(mayor=employee):
        ws1.append([
            "Распоряжение",
            order.name,
            order.id_doc,
            order.get_group_display(),
            order.get_status_display(),
            "-",  # Для распоряжений нет поля актуальности
            order.release_date.strftime('%d.%m.%Y') if order.release_date else ""
        ])

    # Лист 2: Должен ознакомиться
    ws2 = wb.create_sheet("Должен ознакомиться")
    ws2.append(headers)

    # Документы для ознакомления
    for doc in Doc.objects.filter(responsible_not=employee):
        ws2.append([
            "Документ",
            doc.name,
            doc.id_doc,
            ", ".join([t.type_name for t in doc.type_doc.all()]),
            doc.get_status_display(),
            doc.get_actual_display(),
            doc.date.strftime('%d.%m.%Y') if doc.date else ""
        ])

    # Распоряжения для ознакомления
    for order in Orders.objects.filter(responsible_not=employee):
        ws2.append([
            "Распоряжение",
            order.name,
            order.id_doc,
            order.get_group_display(),
            order.get_status_display(),
            "-",
            order.release_date.strftime('%d.%m.%Y') if order.release_date else ""
        ])

    # Лист 3: Ознакомился
    ws3 = wb.create_sheet("Ознакомился")
    ws3.append(headers)

    # Документы, с которыми ознакомился
    for doc in Doc.objects.filter(responsible=employee):
        ws3.append([
            "Документ",
            doc.name,
            doc.id_doc,
            ", ".join([t.type_name for t in doc.type_doc.all()]),
            doc.get_status_display(),
            doc.get_actual_display(),
            doc.date.strftime('%d.%m.%Y') if doc.date else ""
        ])

    # Распоряжения, с которыми ознакомился
    for order in Orders.objects.filter(responsible=employee):
        ws3.append([
            "Распоряжение",
            order.name,
            order.id_doc,
            order.get_group_display(),
            order.get_status_display(),
            "-",
            order.release_date.strftime('%d.%m.%Y') if order.release_date else ""
        ])

    # Лист 4: Ответственный за действия
    ws4 = wb.create_sheet("Ответственный за действия")
    action_headers = ["Тип", "Документ/Распоряжение", "ID", "Действие", "Срок выполнения", "Статус"]
    ws4.append(action_headers)

    # Обработка процессов, где сотрудник ответственный
    for process in Process.objects.filter(responsible_process=employee):
        # Для документов (через name_doc)
        if process.name_doc:
            ws4.append([
                "Документ",
                process.name_doc,
                "-",  # Нет доступа к ID документа
                ", ".join([a.action_name for a in process.action.all()]),
                process.date_deadline.strftime('%d.%m.%Y') if process.date_deadline else "",
                "Выполнено" if process.status_procces else "Не выполнено"
            ])
        
        # Для распоряжений (ищем через обратную связь)
        related_orders = Orders.objects.filter(process=process)
        for order in related_orders:
            ws4.append([
                "Распоряжение",
                order.name,
                order.id_doc,
                ", ".join([a.action_name for a in process.action.all()]),
                process.date_deadline.strftime('%d.%m.%Y') if process.date_deadline else "",
                "Выполнено" if process.status_procces else "Не выполнено"
            ])

    filename = f"employee_{employee.id_number}_{employee.name}_report.xlsx"
    return create_excel_response(wb, filename)


def generate_full_report(workshop):
    """Генерация полного отчёта."""
    wb = Workbook()

    # Лист 1: Не ознакомившиеся по приказам
    ws1 = wb.active
    ws1.title = "Не ознакомившиеся сотрудники по приказам"
    ws1.append(
        ["ФИО сотрудника", "Документ", "ID документа", "Цех", "Участок", "В отпуске"]
    )

    docs = Doc.objects.filter(workshop=workshop)
    for doc in docs:
        if doc.responsible_not.exists():
            for employee in doc.responsible_not.all():
                ws1.append(
                    [
                        employee.name,
                        doc.name,
                        doc.id_doc,
                        employee.workshop.workshop_name if employee.workshop else "",
                        employee.section.section if employee.section else "",
                        employee.on_leave if employee.on_leave else "Нет",
                    ]
                )

    # Лист 2: Не ознакомившиеся по распоряжениям
    ws2 = wb.create_sheet("По распоряжениям")
    ws2.append(
        [
            "ФИО сотрудника",
            "Распоряжение",
            "ID распоряжения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    orders = Orders.objects.filter(workshop=workshop)
    for order in orders:
        if order.responsible_not.exists():
            for employee in order.responsible_not.all():
                ws2.append(
                    [
                        employee.name,
                        order.name,
                        order.id_doc,
                        employee.workshop.workshop_name if employee.workshop else "",
                        employee.section.section if employee.section else "",
                        employee.on_leave if employee.on_leave else "Нет",
                    ]
                )

    # Лист 3: Невыполненные действия по приказам
    ws3 = wb.create_sheet("Невыполненные действия по приказам")
    ws3.append(
        [
            "ФИО сотрудника",
            "Действие",
            "Документ",
            "ID документа",
            "Срок выполнения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    docs = Doc.objects.filter(workshop=workshop)
    for doc in docs:
        processes = doc.process.filter(status_procces=False)
        for process in processes:
            for proc in process.action.all():
                for action in process.responsible_process.all():
                    ws3.append(
                        [
                            action.name,
                            proc.action_name,
                            doc.name,
                            doc.id_doc,
                            (
                                process.date_deadline.strftime("%d.%m.%Y")
                                if process.date_deadline
                                else ""
                            ),
                            action.workshop.workshop_name if action.workshop else "",
                            action.section.section if action.section else "",
                            action.on_leave if action.on_leave else "Нет",
                        ]
                    )

    # Лист 4: Невыполненные действия по распоряжениям
    ws4 = wb.create_sheet("Невыполненные действия по распоряжениям")
    ws4.append(
        [
            "Распоряжение",
            "ID распоряжения",
            "Действие",
            "Ответственный за процесс",
            "Срок выполнения",
            "Цех",
            "Участок",
            "В отпуске",
        ]
    )

    orders = Orders.objects.filter(workshop=workshop)
    for order in orders:
        processes = order.process.filter(status_procces=False)
        for process in processes:
            for action in process.action.all():
                for responsible in process.responsible_process.all():
                    ws4.append(
                        [
                            order.name,
                            order.id_doc,
                            action.action_name,
                            responsible.name,
                            (
                                process.date_deadline.strftime("%d.%m.%Y")
                                if process.date_deadline
                                else ""
                            ),
                            (
                                responsible.workshop.workshop_name
                                if responsible.workshop
                                else ""
                            ),
                            responsible.section.section if responsible.section else "",
                            responsible.on_leave if responsible.on_leave else "Нет",
                        ]
                    )

    return create_excel_response(wb, "full_report.xlsx")


def create_excel_response(workbook, filename):
    """Создание HTTP-ответа с Excel-файлом."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    workbook.save(response)
    return response


def edit_doc(request, id):
    doc = get_object_or_404(Doc, id=id)
    if request.method == "POST":
        form = DocForm(request.POST, request.FILES, instance=doc)
        if form.is_valid():
            form.save()
            return redirect("doc_show_info", id=doc.id)
    else:
        form = DocForm(instance=doc)
    return render(request, "autozavod_app/edit_doc.html", {"form": form, "doc": doc})


def edit_order(request, id):
    order = get_object_or_404(Orders, id=id)
    if request.method == "POST":
        form = OrdersForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            return redirect("order_detail", id=order.id)
    else:
        form = DocForm(instance=order)
    return render(request, "autozavod_app/edit_doc.html", {"form": form, "order": order})

def get_employees_by_filters(request):
    try:
        workshop_ids = request.GET.get('workshop_ids', '').split(',')
        evp_class_ids = request.GET.get('evp_class_ids', '').split(',')
        employee_ids = request.GET.get('employee_ids', '').split(',')
        full_info = request.GET.get('full_info', 'false') == 'true'
        
        # Фильтрация и преобразование ID
        workshop_ids = [int(id) for id in workshop_ids if id.isdigit()]
        evp_class_ids = [int(id) for id in evp_class_ids if id.isdigit()]
        employee_ids = [int(id) for id in employee_ids if id.isdigit()]
        
        # Если запрашиваются конкретные employee_ids
        if employee_ids:
            employees = Stuff.objects.filter(id__in=employee_ids)
        else:
            # Основной запрос по цехам и группам
            if not workshop_ids or not evp_class_ids:
                return JsonResponse({'employees': [], 'count': 0})
            
            employees = Stuff.objects.filter(
                Q(workshop__id__in=workshop_ids) & 
                Q(evp_class__id__in=evp_class_ids) &
                Q(status=True) &
                (Q(on_leave__isnull=True) | Q(on_leave=''))
            ).distinct()
        
        if full_info:
            employees_info = list(employees.values('id', 'name', 'position'))
            return JsonResponse({
                'employees': list(employees.values_list('id', flat=True)),
                'employees_info': employees_info,
                'count': len(employees_info)
            })
        else:
            return JsonResponse({
                'employees': list(employees.values_list('id', flat=True)),
                'count': employees.count()
            })
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'employees': [],
            'count': 0
        }, status=500)

